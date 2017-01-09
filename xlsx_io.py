from openpyxl import Workbook
from openpyxl.styles import PatternFill

class xlsx_io:
    def __init__(self, name):
        self.wb = Workbook()     # Create a new workbook
        self.ws = self.wb.active # Get the active worksheet
        self.columnIDs = {}      # Keep track of which column name is where
        self.workingRow = 1      # Keep track of the row we're writing in

    # text receives a list of strings and writes them into seperate cells in a single row    
    def rawWrite(self, text):
        col = 1
        for t in text:
            print(t)
            self.ws.cell(row=self.workingRow, column=col, value=t)
            col += 1
        self.workingRow += 1


    # Set the column names that will be used when inserting rows
    def setColumns(self, columns):
        workingColumn = 1   # Where we start naming columns
        for col in columns:
            self.columnIDs[col] = workingColumn
            self.ws.cell(row=self.workingRow, column=workingColumn, value=col)
            workingColumn += 1
        self.workingRow += 1

    # Write single rows, the row var is a dict mapping cell values to column names
    def writeRow(self, row):
        for col, val in row.items():
            if(col in self.columnIDs):
                self.ws.cell(row=self.workingRow, column=self.columnIDs[col], value=val)
        self.workingRow += 1

    def fillCells(cells, fill):
        for cell in cells:
            cell.fill = fill

    def generateXLSX(self):
        # Save the file
        self.wb.save("Program.xlsx")
